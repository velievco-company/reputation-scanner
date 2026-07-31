"""
report_generator.py
Создание Excel и PDF отчётов на основе результатов анализа.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)


RISK_COLORS = {
    "LOW_RISK": "C6EFCE",
    "MEDIUM_RISK": "FFEB9C",
    "HIGH_RISK": "FFC7CE",
}


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

def generate_excel_report(
    company_name: str,
    website: str,
    score_summary: dict,
    ai_summary: dict,
) -> bytes:
    """Генерирует .xlsx отчёт в памяти, возвращает bytes для скачивания."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reputation Report"

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    label_font = Font(bold=True)
    risk_level = score_summary.get("risk_level", "MEDIUM_RISK")
    risk_fill = PatternFill(
        start_color=RISK_COLORS.get(risk_level, "FFEB9C"),
        end_color=RISK_COLORS.get(risk_level, "FFEB9C"),
        fill_type="solid",
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reputation Report — {company_name}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Дата анализа:"
    ws["A3"].font = label_font
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws["A4"] = "Сайт:"
    ws["A4"].font = label_font
    ws["B4"] = website or "—"

    ws["A6"] = "Итоговый Reputation Score:"
    ws["A6"].font = label_font
    ws["B6"] = f"{score_summary.get('final_score')}/100"
    ws["B6"].font = Font(bold=True, size=12)

    ws["A7"] = "Уровень риска:"
    ws["A7"].font = label_font
    ws["B7"] = risk_level
    ws["B7"].fill = risk_fill
    ws["B7"].font = Font(bold=True)

    if score_summary.get("confidence_warning"):
        ws["A8"] = score_summary["confidence_warning"]
        ws["A8"].font = Font(italic=True, color="9C6500")
        ws.merge_cells("A8:D8")

    # Детализация по категориям
    ws["A10"] = "Категория"
    ws["B10"] = "Балл"
    ws["C10"] = "Вес"
    ws["D10"] = "Детали"
    for cell in ("A10", "B10", "C10", "D10"):
        ws[cell].font = label_font
        ws[cell].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    breakdown = score_summary.get("breakdown", {})
    weights = score_summary.get("weights_used", {})
    rows = [
        ("Отзывы (Reviews)", breakdown.get("reviews", {}), weights.get("reviews", 0)),
        ("Видимость (Visibility)", breakdown.get("visibility", {}), weights.get("visibility", 0)),
        ("Негатив (Negative)", breakdown.get("negative", {}), weights.get("negative", 0)),
    ]

    row_idx = 11
    for label, data, weight in rows:
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = data.get("score", "—")
        ws[f"C{row_idx}"] = f"{int(weight * 100)}%"
        ws[f"D{row_idx}"] = data.get("detail", "")
        row_idx += 1

    # Найденные риски
    row_idx += 2
    ws[f"A{row_idx}"] = "Найденные риски"
    ws[f"A{row_idx}"].font = Font(bold=True, size=12)
    row_idx += 1

    ws[f"A{row_idx}"] = "Заголовок"
    ws[f"B{row_idx}"] = "Дней назад"
    ws[f"C{row_idx}"] = "Штраф"
    ws[f"D{row_idx}"] = "Ссылка"
    for col in ("A", "B", "C", "D"):
        ws[f"{col}{row_idx}"].font = label_font
    row_idx += 1

    findings = breakdown.get("negative", {}).get("findings", [])
    if not findings:
        ws[f"A{row_idx}"] = "Негативных упоминаний не найдено."
        row_idx += 1
    else:
        for finding in findings:
            ws[f"A{row_idx}"] = finding.get("title", "")
            ws[f"B{row_idx}"] = finding.get("days_ago", "")
            ws[f"C{row_idx}"] = finding.get("penalty_applied", "")
            ws[f"D{row_idx}"] = finding.get("url", "")
            row_idx += 1

    # AI-резюме
    row_idx += 2
    ws[f"A{row_idx}"] = "AI-резюме и рекомендации"
    ws[f"A{row_idx}"].font = Font(bold=True, size=12)
    row_idx += 1

    ws[f"A{row_idx}"] = ai_summary.get("summary", "")
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws[f"A{row_idx}"].alignment = Alignment(wrap_text=True)
    row_idx += 2

    for rec in ai_summary.get("recommendations", []):
        ws[f"A{row_idx}"] = f"• {rec}"
        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        ws[f"A{row_idx}"].alignment = Alignment(wrap_text=True)
        row_idx += 1

    # Ширина колонок
    widths = {"A": 40, "B": 15, "C": 12, "D": 45}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_pdf_report(
    company_name: str,
    website: str,
    score_summary: dict,
    ai_summary: dict,
) -> bytes:
    """Генерирует .pdf отчёт в памяти, возвращает bytes для скачивания."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleStyle", parent=styles["Title"], fontSize=20, spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "HeadingStyle", parent=styles["Heading2"], fontSize=13, spaceBefore=14, spaceAfter=6,
    )
    body_style = styles["BodyText"]

    risk_level = score_summary.get("risk_level", "MEDIUM_RISK")
    risk_color_map = {
        "LOW_RISK": colors.HexColor("#2E7D32"),
        "MEDIUM_RISK": colors.HexColor("#F9A825"),
        "HIGH_RISK": colors.HexColor("#C62828"),
    }

    elements = [
        Paragraph(f"Reputation Report — {company_name}", title_style),
        Paragraph(f"Дата анализа: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style),
        Paragraph(f"Сайт: {website or '—'}", body_style),
        Spacer(1, 0.5 * cm),
    ]

    score_table_data = [
        ["Итоговый Score", f"{score_summary.get('final_score')}/100"],
        ["Уровень риска", risk_level],
    ]
    score_table = Table(score_table_data, colWidths=[8 * cm, 8 * cm])
    score_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (1, 1), (1, 1), risk_color_map.get(risk_level, colors.grey)),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.white),
        ("FONTNAME", (1, 1), (1, 1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(score_table)

    if score_summary.get("confidence_warning"):
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph(
            f'<font color="#9C6500">{score_summary["confidence_warning"]}</font>', body_style
        ))

    elements.append(Paragraph("Детализация по категориям", heading_style))
    breakdown = score_summary.get("breakdown", {})
    weights = score_summary.get("weights_used", {})
    breakdown_data = [["Категория", "Балл", "Вес"]]
    for key, label in [("reviews", "Отзывы"), ("visibility", "Видимость"), ("negative", "Негатив")]:
        data = breakdown.get(key, {})
        breakdown_data.append([label, f"{data.get('score', '—')}/100", f"{int(weights.get(key, 0) * 100)}%"])

    breakdown_table = Table(breakdown_data, colWidths=[6 * cm, 5 * cm, 5 * cm])
    breakdown_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9D9D9")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(breakdown_table)

    elements.append(Paragraph("Найденные риски", heading_style))
    findings = breakdown.get("negative", {}).get("findings", [])
    if not findings:
        elements.append(Paragraph("Негативных упоминаний не найдено.", body_style))
    else:
        for finding in findings:
            elements.append(Paragraph(
                f"• <b>{finding.get('title', '')}</b> ({finding.get('days_ago', '')} дней назад, "
                f"штраф {finding.get('penalty_applied', '')}) — {finding.get('url', '')}",
                body_style,
            ))

    elements.append(Paragraph("AI-резюме", heading_style))
    elements.append(Paragraph(ai_summary.get("summary", ""), body_style))

    elements.append(Paragraph("Рекомендации", heading_style))
    for rec in ai_summary.get("recommendations", []):
        elements.append(Paragraph(f"• {rec}", body_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
